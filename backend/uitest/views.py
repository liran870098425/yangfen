import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.decorators import action
from rest_framework.response import Response

from dvadmin.utils.json_response import SuccessResponse
from dvadmin.utils.viewset import CustomModelViewSet
from uitest.models import BusinessRecord
from uitest.serializers import (
    BusinessRecordSerializer,
    BusinessRecordListSerializer,
)


class BusinessRecordViewSet(CustomModelViewSet):
    """业务数据登记接口"""
    queryset = BusinessRecord.objects.all()
    serializer_class = BusinessRecordSerializer
    list_serializer_class = BusinessRecordListSerializer
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filterset_fields = ('wechat_id', 'is_paid', 'collision_platform', 'shipment_platform', 'record_date')
    search_fields = ('wechat_id', 'demand', 'contact', 'collision_platform', 'shipment_platform')
    ordering_fields = ('record_date', 'created_at', 'price_actual', 'profit')

    # 统计和按月汇总接口不需要权限
    def get_permissions(self):
        if self.action in ['stats', 'monthly_list', 'export_excel']:
            return []
        return super().get_permissions()

    def destroy(self, request, *args, **kwargs):
        """删除后重新计算所有汇总"""
        instance = self.get_object()
        instance.delete()
        BusinessRecord.update_all_totals()
        return SuccessResponse(msg="删除成功")

    @action(methods=['get'], detail=False)
    def stats(self, request):
        """获取统计数据"""
        from django.db.models import Sum
        total_stats = BusinessRecord.objects.aggregate(
            total_profit_all=Sum('profit'),
            total_platform_deduct=Sum('platform_deduct'),
            total_cost_all=Sum('cost_deduct'),
            monthly_profit_total=Sum('profit'),
            monthly_platform_deduct=Sum('platform_deduct'),
            monthly_cost_total=Sum('cost_deduct'),
        )
        print('=== STATS DEBUG ===', total_stats)
        return SuccessResponse(data=total_stats)

    @action(methods=['get'], detail=False)
    def monthly_list(self, request):
        """获取按月汇总列表"""
        from django.db.models import Sum, Count
        from django.db.models.functions import TruncMonth

        qs = BusinessRecord.objects.annotate(
            month=TruncMonth('record_date')
        ).values('month').annotate(
            total_platform_deduct=Sum('platform_deduct'),
            total_cost_deduct=Sum('cost_deduct'),
            total_profit=Sum('profit'),
            total_receipt=Sum('monthly_receipt'),
            count=Count('id'),
        ).order_by('-month')

        return SuccessResponse(data=list(qs))

    @action(methods=['get'], detail=False)
    def export_excel(self, request):
        """导出 Excel"""
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "业务数据"

        headers = [
            '微信号', '需求', '数量', '联系方式', '撞车平台', '出货平台',
            '付款方式', '是否付款', '应付价格', '实付价格', '平台扣除',
            '成本扣除', '红包', '盈余', '月实收(流水)', '月平台总扣除',
            '月成本总', '月总盈余', '平台总扣除', '成本总', '总盈余',
            '业务日期', '备注'
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.wechat_id or '')
            ws.cell(row=row_num, column=2, value=record.demand or '')
            ws.cell(row=row_num, column=3, value=record.quantity or '')
            ws.cell(row=row_num, column=4, value=record.contact or '')
            ws.cell(row=row_num, column=5, value=record.collision_platform or '')
            ws.cell(row=row_num, column=6, value=record.shipment_platform or '')
            ws.cell(row=row_num, column=7, value=record.payment_method or '')
            ws.cell(row=row_num, column=8, value='是' if record.is_paid else '否')
            ws.cell(row=row_num, column=9, value=record.price_payable or 0)
            ws.cell(row=row_num, column=10, value=record.price_actual or 0)
            ws.cell(row=row_num, column=11, value=record.platform_deduct or 0)
            ws.cell(row=row_num, column=12, value=record.cost_deduct or 0)
            ws.cell(row=row_num, column=13, value=record.redpacket or 0)
            ws.cell(row=row_num, column=14, value=record.profit or 0)
            ws.cell(row=row_num, column=15, value=record.monthly_receipt or 0)
            ws.cell(row=row_num, column=16, value=record.monthly_platform_deduct or 0)
            ws.cell(row=row_num, column=17, value=record.monthly_cost_total or 0)
            ws.cell(row=row_num, column=18, value=record.monthly_profit_total or 0)
            ws.cell(row=row_num, column=19, value=record.total_platform_deduct or 0)
            ws.cell(row=row_num, column=20, value=record.total_cost_all or 0)
            ws.cell(row=row_num, column=21, value=record.total_profit_all or 0)
            ws.cell(row=row_num, column=22, value=str(record.record_date) if record.record_date else '')
            ws.cell(row=row_num, column=23, value=record.remark or '')

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['W'].width = 30

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="business_records.xlsx"'
        wb.save(response)
        return response
