import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin.utils.json_response import SuccessResponse, ErrorResponse
from escortprice.models import EscortPrice
from escortprice.serializers import (
    EscortPriceSerializer,
    EscortPriceListSerializer,
)


class EscortPriceViewSet(CustomModelViewSet):
    """三角洲护航价格管理接口"""
    queryset = EscortPrice.objects.all()
    serializer_class = EscortPriceSerializer
    list_serializer_class = EscortPriceListSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    search_fields = ('package_name', 'remark')
    ordering_fields = ('package_type', 'sell_price', 'created_at')

    def get_permissions(self):
        return []

    @action(methods=['get'], detail=False)
    def export_excel(self, request):
        """导出Excel价格表"""
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "护航价格表"

        headers = ['序号', '套餐类型', '套餐名称', '出售价格(R)', '保底游戏币(w)', '炸单额外加(w)', '对局场次', '备注', '创建时间']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, item in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=item.package_type or '')
            ws.cell(row=row_num, column=3, value=item.package_name or '')
            ws.cell(row=row_num, column=4, value=item.sell_price or 0)
            ws.cell(row=row_num, column=5, value=item.guarantee_money or 0)
            ws.cell(row=row_num, column=6, value=item.bomb_extra or 0)
            ws.cell(row=row_num, column=7, value=item.game_times or '')
            ws.cell(row=row_num, column=8, value=item.remark or '')
            ws.cell(row=row_num, column=9, value=str(item.created_at) if item.created_at else '')

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="escort_prices.xlsx"'
        wb.save(response)
        return response

    @action(methods=['post'], detail=False)
    def import_excel(self, request):
        """批量导入Excel"""
        if 'file' not in request.FILES:
            return ErrorResponse(msg="请上传Excel文件")

        excel_file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            success_count = 0
            error_list = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[1]:
                    continue

                try:
                    package_type = str(row[1]).strip() if row[1] else ''
                    if not package_type or package_type not in dict(EscortPrice.PACKAGE_TYPE_CHOICES):
                        error_list.append(f"第{row_num}行: 套餐类型无效，必须是：体验单/绝密监狱航天/绝密巴克什/赌约单")
                        continue

                    package_name = str(row[2]).strip() if row[2] else ''
                    if not package_name:
                        error_list.append(f"第{row_num}行: 套餐名称不能为空")
                        continue

                    sell_price = float(row[3]) if row[3] else 0
                    guarantee_money = float(row[4]) if row[4] else 0
                    bomb_extra = float(row[5]) if row[5] else 0
                    game_times = str(row[6]).strip() if row[6] else ''
                    remark = str(row[7]).strip() if row[7] else ''

                    EscortPrice.objects.update_or_create(
                        package_type=package_type,
                        package_name=package_name,
                        defaults={
                            'sell_price': sell_price,
                            'guarantee_money': guarantee_money,
                            'bomb_extra': bomb_extra,
                            'game_times': game_times,
                            'remark': remark,
                        }
                    )
                    success_count += 1
                except Exception as e:
                    error_list.append(f"第{row_num}行: {str(e)}")

            msg = f"成功导入 {success_count} 条数据"
            return SuccessResponse(msg=msg, data={'success': success_count, 'errors': error_list})
        except Exception as e:
            return ErrorResponse(msg=f"导入失败: {str(e)}")
