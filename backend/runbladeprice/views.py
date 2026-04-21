import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin.utils.json_response import SuccessResponse, ErrorResponse
from runbladeprice.models import RunBladePrice
from runbladeprice.serializers import (
    RunBladePriceSerializer,
    RunBladePriceListSerializer,
)


class RunBladePriceViewSet(CustomModelViewSet):
    """三角洲跑刀跑手价格管理接口"""
    queryset = RunBladePrice.objects.all()
    serializer_class = RunBladePriceSerializer
    list_serializer_class = RunBladePriceListSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    search_fields = ('remark',)
    ordering_fields = ('business_type', 'grid_type', 'rmb', 'create_datetime')

    def get_permissions(self):
        return []

    @action(methods=['get'], detail=False)
    def export_excel(self, request):
        """导出Excel价格表"""
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "跑刀跑手价格表"

        headers = ['序号', '业务类型', '物品格子', '单人单价(R)', '人民币R', '对应游戏币', '备注说明', '创建时间']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, item in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=item.business_type or '')
            ws.cell(row=row_num, column=3, value=item.grid_type or '')
            ws.cell(row=row_num, column=4, value=item.unit_price or '')
            ws.cell(row=row_num, column=5, value=item.rmb or 0)
            ws.cell(row=row_num, column=6, value=item.game_money or '')
            ws.cell(row=row_num, column=7, value=item.remark or '')
            ws.cell(row=row_num, column=8, value=str(item.create_datetime) if item.create_datetime else '')

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="runblade_prices.xlsx"'
        wb.save(response)
        return response

    @action(methods=['post'], detail=False)
    def import_excel(self, request):
        """批量导入Excel"""
        if 'file' not in request.FILES:
            return ErrorResponse(msg="请上传Excel文件")

        excel_file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            success_count = 0
            error_list = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[1]:
                    continue

                try:
                    business_type = str(row[1]).strip() if row[1] else ''
                    if not business_type or business_type not in dict(RunBladePrice.TYPE_CHOICES):
                        error_list.append(f"第{row_num}行: 业务类型无效，必须是：跑刀/跑手")
                        continue

                    grid_type = str(row[2]).strip() if row[2] else ''
                    if not grid_type or grid_type not in dict(RunBladePrice.GRID_CHOICES):
                        error_list.append(f"第{row_num}行: 物品格子无效，必须是：9格/6格/4格/2格/4-6格")
                        continue

                    unit_price = str(row[3]).strip() if row[3] else ''
                    rmb = float(row[4]) if row[4] else 0
                    game_money = str(row[5]).strip() if row[5] else ''
                    remark = str(row[6]).strip() if row[6] else ''

                    RunBladePrice.objects.update_or_create(
                        business_type=business_type,
                        grid_type=grid_type,
                        defaults={
                            'unit_price': unit_price,
                            'rmb': rmb,
                            'game_money': game_money,
                            'remark': remark,
                        }
                    )
                    success_count += 1
                except Exception as e:
                    error_list.append(f"第{row_num}行: {str(e)}")

            msg = f"成功导入 {success_count} 条数据"
            return SuccessResponse(msg=msg, data={'success': success_count, 'errors': error_list})
        except Exception as e:
            return ErrorResponse(msg=f"导入失败: {str(e)}")
