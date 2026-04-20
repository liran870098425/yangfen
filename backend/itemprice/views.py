import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin.utils.json_response import SuccessResponse, ErrorResponse
from itemprice.models import ItemPrice
from itemprice.serializers import (
    ItemPriceSerializer,
    ItemPriceListSerializer,
)


class ItemPriceViewSet(CustomModelViewSet):
    """物品价格管理接口"""
    queryset = ItemPrice.objects.all()
    serializer_class = ItemPriceSerializer
    list_serializer_class = ItemPriceListSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    search_fields = ('item_name', 'item_remark')
    ordering_fields = ('item_name', 'created_at', 'out_price', 'reference_price')

    def get_permissions(self):
        return []

    # ---------------------- 以下代码保持不变 ----------------------
    @action(methods=['get'], detail=False)
    def export_excel(self, request):
        """导出Excel价格表"""
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "物品价格表"

        headers = ['序号', '物资名称', '参考价格', '出价格', '物品备注规格', '创建时间', '更新时间']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, item in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=item.item_name or '')
            ws.cell(row=row_num, column=3, value=item.reference_price or 0)
            ws.cell(row=row_num, column=4, value=item.out_price or 0)
            ws.cell(row=row_num, column=5, value=item.item_remark or '')
            ws.cell(row=row_num, column=6, value=str(item.created_at) if item.created_at else '')
            ws.cell(row=row_num, column=7, value=str(item.updated_at) if item.updated_at else '')

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="item_prices.xlsx"'
        wb.save(response)
        return response

    @action(methods=['post'], detail=False)
    def import_excel(self, request):
        """批量导入Excel"""
        if 'file' not in request.FILES:
            return ErrorResponse(msg="请上传Excel文件")

        excel_file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            success_count = 0
            error_list = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]:
                    continue

                try:
                    item_name = str(row[0]).strip() if row[0] else ''
                    if not item_name:
                        continue

                    reference_price = float(row[1]) if row[1] else 0
                    out_price = float(row[2]) if row[2] else 0
                    item_remark = str(row[3]).strip() if row[3] else ''

                    ItemPrice.objects.update_or_create(
                        item_name=item_name,
                        defaults={
                            'reference_price': reference_price,
                            'out_price': out_price,
                            'item_remark': item_remark,
                        }
                    )
                    success_count += 1
                except Exception as e:
                    error_list.append(f"第{row_num}行: {str(e)}")

            msg = f"成功导入 {success_count} 条数据"
            return SuccessResponse(msg=msg, data={'success': success_count, 'errors': error_list})
        except Exception as e:
            return ErrorResponse(msg=f"导入失败: {str(e)}")

    @action(methods=['post'], detail=False)
    def batch_update_price(self, request):
        """批量修改价格"""
        data = request.data
        if not isinstance(data, list):
            return ErrorResponse(msg="数据格式错误，应为列表")

        success_count = 0
        error_list = []

        for idx, item in enumerate(data):
            try:
                item_id = item.get('id')
                if not item_id:
                    error_list.append(f"第{idx + 1}条: 缺少id字段")
                    continue

                price_obj = ItemPrice.objects.filter(pk=item_id).first()
                if not price_obj:
                    error_list.append(f"第{idx + 1}条: 物品不存在")
                    continue

                if 'out_price' in item:
                    price_obj.out_price = float(item['out_price'])
                if 'reference_price' in item:
                    price_obj.reference_price = float(item['reference_price'])
                if 'item_remark' in item:
                    price_obj.item_remark = item['item_remark']

                price_obj.save()
                success_count += 1
            except Exception as e:
                error_list.append(f"第{idx + 1}条: {str(e)}")

        msg = f"成功更新 {success_count} 条数据"
        return SuccessResponse(msg=msg, data={'success': success_count, 'errors': error_list})